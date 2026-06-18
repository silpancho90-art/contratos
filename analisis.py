# -*- coding: utf-8 -*-
"""
Analisis minucioso persona por persona:
- Identifica FALTA CONTRATO en Control de Contratos BNB
- Recupera TODOS los registros de cada persona en Estructura de Personal V.2
  (cruce por Carnet + validacion/recuperacion por Nombre y Apellidos)
- Orden cronologico, estado/cargo/ciudad/supervisor actual (registro mas reciente)
- Regla especial: si hay registro mas reciente ACTIVO => ACTIVO
- Detecta inconsistencias
- Genera Validacion_Falta_Contratos_BNB.xlsx con 3 hojas
"""
import re
import unicodedata
import pandas as pd
from datetime import datetime

CTRL_FN = "CONTROL DE CONTRATOS POR PROYECTOS BNB.xlsx"
EST_FN = "ESTRUCTURA DE PERSONAL V.2.xlsx"
OUT_FN = "Validacion_Falta_Contratos_BNB.xlsx"
LOG_FN = "trazabilidad_log.txt"

CARGO_RANK = {
    "SUPERVISOR REGIONAL": 5,
    "SUPERVISOR": 4,
    "AFILIADOR": 3,
    "ACTIVADOR": 2,
    "CAPACITACION": 1,
    "PASANTE": 0,
}

# ---------------- Helpers ----------------

def strip_accents(s):
    return "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))

def norm_text(s):
    if s is None or (isinstance(s, float) and pd.isna(s)):
        return ""
    s = str(s).strip().upper()
    s = strip_accents(s)
    s = re.sub(r"[^A-Z0-9 ]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s

def clean_ci(v):
    """Normaliza carnet a string sin decimales."""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ""
    if isinstance(v, float):
        if pd.isna(v):
            return ""
        return str(int(v))
    s = str(v).strip()
    if s.lower() in ("nan", "none", ""):
        return ""
    # quitar .0 final
    if re.fullmatch(r"\d+\.0", s):
        s = s[:-2]
    return s

def ci_valido(ci):
    return ci not in ("", "0")

def name_tokens(s):
    return set(t for t in norm_text(s).split() if len(t) > 1)

def names_match(set_a, set_b):
    """True si los nombres coinciden de forma robusta (uno subconjunto del otro)."""
    if not set_a or not set_b:
        return False
    small, big = (set_a, set_b) if len(set_a) <= len(set_b) else (set_b, set_a)
    if small.issubset(big):
        return True
    inter = set_a & set_b
    union = set_a | set_b
    return len(inter) / len(union) >= 0.6 if union else False

def names_match_strong(set_a, set_b):
    """Coincidencia fuerte para recuperar por nombre cuando no hay carnet valido."""
    if len(set_a) < 2 or len(set_b) < 2:
        return False
    small, big = (set_a, set_b) if len(set_a) <= len(set_b) else (set_b, set_a)
    return small.issubset(big) and len(small) >= 2

def fmt_date(d):
    if d is None or pd.isna(d):
        return ""
    if isinstance(d, (pd.Timestamp, datetime)):
        return d.strftime("%d/%m/%Y")
    return str(d)

def cargo_rank(c):
    return CARGO_RANK.get(norm_text(c), -1)

# ---------------- Carga ----------------

ctrl = pd.read_excel(CTRL_FN, sheet_name="CONTRATOS BNB", header=0, dtype=str)
est = pd.read_excel(EST_FN, sheet_name="estructura de personal", header=0)

ctrl.columns = [str(c).strip() for c in ctrl.columns]
est.columns = [str(c).strip() for c in est.columns]

EST_NOMBRE = "Nombres y Apellidos"
EST_FECHA = "Fecha de ingreso"
EST_CARGO = "Cargo"
EST_CIUDAD = "Ciudad"
EST_SUP = "SUPERVISOR"
EST_ESTADO = "Estado (Activo/Inactivo)"
EST_CI = "C.I."
EST_CODIGO = "CODIGO"

# Pre-procesar estructura
est = est.copy()
est["_CI"] = est[EST_CI].map(clean_ci)
est["_NAME"] = est[EST_NOMBRE].astype(str)
est["_NTOK"] = est["_NAME"].map(name_tokens)
est["_FECHA"] = pd.to_datetime(est[EST_FECHA], errors="coerce")
est["_ESTADO"] = est[EST_ESTADO].map(norm_text)
est["_CARGO"] = est[EST_CARGO].astype(str).str.strip()

# ---------------- FALTA CONTRATO ----------------
ctrl["_EC"] = ctrl["ESTADO DE CONTRATO"].map(norm_text)
falta = ctrl[ctrl["_EC"] == "FALTA CONTRATO"].copy()
falta["_CI"] = falta["C.I."].map(clean_ci)

# Construir nombre completo del control
def ctrl_fullname(row):
    nc = str(row.get("NOMBRE COMPLETO", "") or "").strip()
    if nc and nc.lower() != "nan":
        return nc
    parts = [str(row.get(c, "") or "").strip() for c in ["NOMBRES", "APELLIDO PATERNO", "APELLIDO MATERNO"]]
    return " ".join(p for p in parts if p and p.lower() != "nan")

falta["_NAME"] = falta.apply(ctrl_fullname, axis=1)
falta["_NTOK"] = falta["_NAME"].map(name_tokens)

# Agrupar en personas unicas: por carnet valido, si no por nombre normalizado
def person_key(row):
    if ci_valido(row["_CI"]):
        return ("CI", row["_CI"])
    return ("NM", norm_text(row["_NAME"]))

falta["_PKEY"] = falta.apply(person_key, axis=1)

# ---------------- Procesar cada persona ----------------
resumen_rows = []
traza_rows = []
mov_rows = []
log_lines = []

personas = []
seen = set()
for _, r in falta.iterrows():
    k = r["_PKEY"]
    if k in seen:
        continue
    seen.add(k)
    personas.append(k)

print(f"Personas unicas con FALTA CONTRATO: {len(personas)}")
print(f"Filas FALTA CONTRATO: {len(falta)}")

total_traza = 0
for k in personas:
    grp = falta[falta["_PKEY"] == k]
    # datos representativos del control (tomar el de carnet valido / primero)
    base = grp.iloc[0]
    carnet = base["_CI"]
    nombre_ctrl = base["_NAME"]
    ntok_ctrl = base["_NTOK"]
    codigos_ctrl = sorted(set(str(x).strip() for x in grp["CODIGO"].tolist() if str(x).strip().lower() != "nan"))

    # ---- Recuperar registros en estructura ----
    matched_idx = {}  # idx -> metodo
    if ci_valido(carnet):
        for idx, er in est[est["_CI"] == carnet].iterrows():
            matched_idx[idx] = "CARNET"
    # recuperar/validar por nombre
    for idx, er in est.iterrows():
        if idx in matched_idx:
            continue
        if ci_valido(carnet):
            # añadir por nombre solo si coincidencia fuerte (posible carnet distinto/typo)
            if names_match_strong(ntok_ctrl, er["_NTOK"]):
                matched_idx[idx] = "NOMBRE"
        else:
            # sin carnet valido: recuperar por nombre fuerte
            if names_match_strong(ntok_ctrl, er["_NTOK"]):
                matched_idx[idx] = "NOMBRE"

    recs = est.loc[list(matched_idx.keys())].copy() if matched_idx else est.iloc[0:0].copy()
    if len(recs):
        recs["_METODO"] = [matched_idx[i] for i in recs.index]
        # validacion de nombre por registro
        recs["_NAMEOK"] = recs["_NTOK"].map(lambda t: "SI" if names_match(ntok_ctrl, t) else "NO")
        # orden cronologico (fecha asc); NaT al final
        recs = recs.sort_values(by=["_FECHA"], na_position="last", kind="mergesort").reset_index(drop=True)

    n_recs = len(recs)
    total_traza += n_recs

    # ---- Determinar registro vigente / estado actual ----
    estado_actual = cargo_actual = ciudad_actual = sup_actual = fecha_ult = ""
    conclusion = ""
    vigente_pos = -1
    conflicto = ""

    if n_recs == 0:
        conclusion = "NO ENCONTRADO EN ESTRUCTURA"
    else:
        # fecha maxima (registros con fecha)
        with_date = recs[recs["_FECHA"].notna()]
        if len(with_date):
            max_date = with_date["_FECHA"].max()
            cand = with_date[with_date["_FECHA"] == max_date]
        else:
            # sin fechas: usar todos como candidatos
            max_date = None
            cand = recs.copy()
        # elegir vigente: preferir ACTIVO, luego mayor rango de cargo
        cand = cand.copy()
        cand["_isACT"] = (cand["_ESTADO"] == "ACTIVO").astype(int)
        cand["_rank"] = cand["_CARGO"].map(cargo_rank)
        cand = cand.sort_values(by=["_isACT", "_rank"], ascending=[False, False], kind="mergesort")
        vigente = cand.iloc[0]
        # localizar posicion del vigente en recs (por igualdad de fila)
        vigente_pos = vigente.name if vigente.name in recs.index else None
        # identificar indice posicional
        # marcar vigente por coincidencia exacta de la fila original
        vig_orig_idx = vigente.name

        estado_actual = vigente["_ESTADO"]
        cargo_actual = vigente["_CARGO"]
        ciudad_actual = str(vigente[EST_CIUDAD]).strip()
        sup_actual = str(vigente[EST_SUP]).strip()
        fecha_ult = fmt_date(vigente["_FECHA"])

        # Regla especial: si en la fecha mas reciente hay algun ACTIVO => ACTIVO
        hay_activo_reciente = (cand["_isACT"].sum() > 0)
        if hay_activo_reciente:
            estado_actual = "ACTIVO"
        # detectar conflicto en la fecha mas reciente
        estados_cand = set(cand["_ESTADO"].tolist())
        if len(estados_cand) > 1:
            conflicto = "Conflicto en fecha mas reciente (" + fmt_date(max_date) + "): estados " + ", ".join(sorted(estados_cand))

        if estado_actual == "ACTIVO":
            conclusion = "ACTIVO - REQUIERE CONTRATO"
        else:
            conclusion = "INACTIVO - NO REQUIERE CONTRATO"

    # ---- Hoja 2: Trazabilidad ----
    for pos, (_, er) in enumerate(recs.iterrows()):
        es_vig = "Si" if (n_recs and er.name == vig_orig_idx) else "No"
        traza_rows.append({
            "Carnet": carnet if ci_valido(carnet) else (er["_CI"] or ""),
            "Nombre": er[EST_NOMBRE],
            "Fecha de Ingreso": fmt_date(er["_FECHA"]),
            "Cargo": er["_CARGO"],
            "Estado": er[EST_ESTADO],
            "Ciudad": er[EST_CIUDAD],
            "Supervisor": er[EST_SUP],
            "Es Registro Vigente (Si/No)": es_vig,
            "Codigo (Estructura)": er[EST_CODIGO],
            "Metodo de Cruce": er["_METODO"],
            "Coincide Nombre": er["_NAMEOK"],
        })

    # ---- Hoja 3: Movimientos ----
    if n_recs:
        primero = recs.iloc[0]
        ultimo_crono = recs[recs["_FECHA"].notna()]
        if len(ultimo_crono):
            ultimo = ultimo_crono.iloc[-1]
        else:
            ultimo = recs.iloc[-1]
        cargos_secuencia = [c for c in recs["_CARGO"].tolist()]
        # secuencia legible de evolucion (cargos distintos consecutivos)
        evol = []
        for c in cargos_secuencia:
            if not evol or evol[-1] != c:
                evol.append(c)
        evol_str = " -> ".join(evol)
        mov_rows.append({
            "Carnet": carnet if ci_valido(carnet) else "",
            "Nombre Completo": primero[EST_NOMBRE],
            "Cargo Inicial": primero["_CARGO"],
            "Fecha Inicial": fmt_date(primero["_FECHA"]),
            "Cargo Final": cargo_actual,
            "Fecha Final": fecha_ult,
            "Estado Final": estado_actual,
            "Cantidad de Cargos Distintos": len(set(cargos_secuencia)),
            "Evolucion Detectada": evol_str,
            "Conflicto": conflicto,
        })
    else:
        mov_rows.append({
            "Carnet": carnet if ci_valido(carnet) else "",
            "Nombre Completo": nombre_ctrl,
            "Cargo Inicial": "",
            "Fecha Inicial": "",
            "Cargo Final": "",
            "Fecha Final": "",
            "Estado Final": "",
            "Cantidad de Cargos Distintos": 0,
            "Evolucion Detectada": "NO ENCONTRADO EN ESTRUCTURA",
            "Conflicto": "",
        })

    # ---- Hoja 1: Resumen ----
    resumen_rows.append({
        "Carnet": carnet if ci_valido(carnet) else (nombre_ctrl and "" or ""),
        "Nombre Completo": nombre_ctrl,
        "Estado de Contrato": "FALTA CONTRATO",
        "Estado Actual": estado_actual,
        "Cargo Actual": cargo_actual,
        "Fecha Ultimo Ingreso": fecha_ult,
        "Ciudad Actual": ciudad_actual,
        "Supervisor Actual": sup_actual,
        "Cantidad de Registros Encontrados": n_recs,
        "Conclusion": conclusion,
        "Codigo(s) Control": ", ".join(codigos_ctrl),
        "Observacion": conflicto,
    })

    # ---- Log detallado ----
    log_lines.append("=" * 90)
    log_lines.append(f"PERSONA: {nombre_ctrl}  | Carnet Control: {carnet or '(sin carnet)'}  | Codigos: {', '.join(codigos_ctrl)}")
    log_lines.append(f"Registros encontrados en Estructura: {n_recs}")
    if n_recs:
        for pos, (_, er) in enumerate(recs.iterrows(), 1):
            vig = " <== VIGENTE" if er.name == vig_orig_idx else ""
            log_lines.append(
                f"  Reg {pos}: Fecha={fmt_date(er['_FECHA'])} | Cargo={er['_CARGO']} | "
                f"Estado={er[EST_ESTADO]} | Ciudad={er[EST_CIUDAD]} | Sup={er[EST_SUP]} | "
                f"Carnet={er['_CI']} | Cruce={er['_METODO']} | NombreOK={er['_NAMEOK']}{vig}"
            )
        log_lines.append(f"  >> Estado Actual: {estado_actual} | Cargo Actual: {cargo_actual} | Fecha vigente: {fecha_ult}")
        if conflicto:
            log_lines.append(f"  >> {conflicto}")
    log_lines.append(f"  >> CONCLUSION: {conclusion}")

# ---------------- Construir DataFrames ----------------
df_resumen = pd.DataFrame(resumen_rows)
df_traza = pd.DataFrame(traza_rows)
df_mov = pd.DataFrame(mov_rows)

# Orden resumen: por conclusion (ACTIVO primero) y nombre
orden_concl = {"ACTIVO - REQUIERE CONTRATO": 0, "INACTIVO - NO REQUIERE CONTRATO": 1, "NO ENCONTRADO EN ESTRUCTURA": 2}
df_resumen["_o"] = df_resumen["Conclusion"].map(lambda x: orden_concl.get(x, 9))
df_resumen = df_resumen.sort_values(by=["_o", "Nombre Completo"]).drop(columns=["_o"]).reset_index(drop=True)

# ---------------- Exportar Excel ----------------
with pd.ExcelWriter(OUT_FN, engine="openpyxl") as xw:
    df_resumen.to_excel(xw, sheet_name="RESUMEN EJECUTIVO", index=False)
    df_traza.to_excel(xw, sheet_name="TRAZABILIDAD COMPLETA", index=False)
    df_mov.to_excel(xw, sheet_name="ANALISIS DE MOVIMIENTOS", index=False)

# Formato basico (ancho de columnas + cabecera)
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

wb = load_workbook(OUT_FN)
header_fill = PatternFill("solid", fgColor="1F4E78")
header_font = Font(bold=True, color="FFFFFF")
for ws in wb.worksheets:
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    ws.freeze_panes = "A2"
    # forzar formato texto en columnas de identificadores (Carnet / Codigo)
    text_cols = set()
    for cell in ws[1]:
        h = str(cell.value or "").lower()
        if "carnet" in h or "codigo" in h:
            text_cols.add(cell.column_letter)
    for col in ws.columns:
        maxlen = 0
        letter = col[0].column_letter
        for c in col:
            v = "" if c.value is None else str(c.value)
            maxlen = max(maxlen, len(v))
            if letter in text_cols:
                c.number_format = "@"
        ws.column_dimensions[letter].width = min(max(maxlen + 2, 10), 45)
wb.save(OUT_FN)

with open(LOG_FN, "w", encoding="utf-8") as f:
    f.write("\n".join(log_lines))

# ---------------- Verificacion ----------------
print("\n=== VERIFICACION ===")
print(f"Personas unicas con FALTA CONTRATO procesadas: {len(personas)}")
print(f"Filas en Hoja RESUMEN: {len(df_resumen)}")
print(f"Filas en Hoja TRAZABILIDAD: {len(df_traza)} (registros de estructura recuperados)")
print(f"Filas en Hoja MOVIMIENTOS: {len(df_mov)}")
print(f"Suma registros encontrados (resumen): {df_resumen['Cantidad de Registros Encontrados'].sum()}")
print("Distribucion de Conclusiones:")
print(df_resumen["Conclusion"].value_counts().to_string())
print(f"\nArchivo generado: {OUT_FN}")
print(f"Log detallado: {LOG_FN}")

# guardar tambien conteos para checks posteriores
df_resumen.to_pickle("_resumen.pkl")
df_traza.to_pickle("_traza.pkl")
